from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font

def create_excel(servico, credores):
    # Create a new Workbook
    wb = Workbook()

    # Get the active sheet
    sheet = wb.active

    # Set headers
    format_cell(sheet['A1'], "Serviço")
    format_cell(sheet['B1'], "Credor")

    # Set cell dimensions 
    sheet.column_dimensions['A'].width = 50
    sheet.row_dimensions[1].height = 30

    # Calculate the ending column letter
    end_credor_letter = get_column_letter(len(credores) + 1)
    sheet.merge_cells(f'B1:{end_credor_letter}1')

    # Total column
    total_column_letter = get_column_letter(len(credores) + 2)
    format_cell(sheet[f'{total_column_letter}1'], "Total")

    # Participants column
    participants_column_letter = get_column_letter(len(credores) + 3)
    end_participants_column_letter = get_column_letter(len(credores) + 3 + len(credores) - 1)
    format_cell(sheet[f'{participants_column_letter}1'], "Participantes")
    sheet.merge_cells(f'{participants_column_letter}1:{end_participants_column_letter}1')

    # Total column
    participants_total_column_letter = get_column_letter(len(credores) + 3 + len(credores))
    format_cell(sheet[f'{participants_total_column_letter}1'], "Total")

    # Balance column
    balance_column_letter = get_column_letter(len(credores) + 3 + len(credores) + 1)
    end_balance_column_letter = get_column_letter(len(credores) + 3 + len(credores) + len(credores))
    format_cell(sheet[f'{balance_column_letter}1'], "Balanço geral")
    sheet.merge_cells(f'{balance_column_letter}1:{end_balance_column_letter}1')

    # Add Credores
    for index, nome in enumerate(credores, start=2):
        cell = sheet.cell(row=2, column=index, value=nome)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[get_column_letter(index)].width = 12

    sum_total_formula=f"=SUM({get_column_letter(len(credores) + 2)}4:{get_column_letter(len(credores) + 2)}10000)"
    sheet.cell(row=3, column=len(credores) + 2, value=sum_total_formula).number_format = 'R$ #,##0.00'

    # Set Credores number format 
    for i in range(0, len(credores) + 1):
        for row in range(4, 50):
            sheet.cell(row=row, column=1 + i, value="").number_format = 'R$ #,##0.00'

    # Sum values for "Total" column
    for row in range(4, 50):
        total_formula = f"=SUM(B{row}:{get_column_letter(len(credores) + 1)}{row})"
        sheet.cell(row=row, column=len(credores) + 2, value=total_formula).number_format = 'R$ #,##0.00'

    # Add Participants
    for index, nome in enumerate(credores, start=len(credores) + 3):
        cell = sheet.cell(row=2, column=index, value=nome)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[get_column_letter(index)].width = 12

    # Sum values for "Total" column
    for row in range(4, 50):
        total_formula = f"=SUM({participants_column_letter}{row}:{end_participants_column_letter}{row})"
        sheet.cell(row=row, column=len(credores) + 3 + len(credores), value=total_formula)

    sum_total_participants_formula=f"=SUM({participants_total_column_letter}4:{participants_total_column_letter}10000)"
    sheet.cell(row=3, column=len(credores) + 3 + len(credores), value=sum_total_participants_formula)

    # Add credores to balance
    for index, nome in enumerate(credores, start=len(credores) + 3 + len(credores) + 1):
        cell = sheet.cell(row=2, column=index, value=nome)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[get_column_letter(index)].width = 12

    # Populate balance
    for i in range (0, len(credores)):
        row_credor = get_column_letter(2 + i)
        row_participant = get_column_letter(len(credores) + 3 + i)
        column_balance = len(credores) + 3 + len(credores) + 1 + i
        sheet.cell(row=3, column=column_balance, value=f"=SUM({get_column_letter(column_balance)}4:{get_column_letter(column_balance)}50)").number_format = 'R$ #,##0.00'

        print(f"row_credor: {row_credor}")
        print(f"row_participant: {row_participant}")
        print(f"column_balance: {column_balance}")

        for row in range(4, 50):
            total_formula = f"=IFS({row_participant}{row} >= 1, {row_credor}{row}-({row_participant}{row}*(${total_column_letter}{row}/${participants_total_column_letter}{row})),{row_participant}{row} = 0, {row_credor}{row})"
            sheet.cell(row=row, column=column_balance, value=total_formula)

    # Set fill color for the rows
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for row in sheet.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.fill = gray_fill

    # Save the workbook to a local file
    wb.save("my_spreadsheet.xlsx")

    print(f"Excel file created with structure: Serviço, Credor, {', '.join(credores)}, Total")


def format_cell(cell, value):
    header_fill = PatternFill(start_color="82cfe8", end_color="82cfe8", fill_type="solid")
    cell.value = value
    cell.fill = header_fill
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Example usage
servico = "Consultoria"
credores = ["John", "Jane", "Alice", "Juquinha"]
create_excel(servico, credores)
